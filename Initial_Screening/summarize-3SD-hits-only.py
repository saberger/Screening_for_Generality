import pandas as pd
from glob import glob
import openpyxl

# Combine evaluation sheets
workbook = openpyxl.Workbook()
sheet = workbook.active
workbook.save('output/alldata_3SD.xlsx')

book_ro = openpyxl.load_workbook('output/alldata_3SD.xlsx')
book = book_ro
sheet1 = book.active

colx = 1
wb = openpyxl.load_workbook('output/alldata_3SD.xlsx')
sheet = wb.active

sheet1['A1'] = "Enzyme ID"
sheet1['B1'] = "Carbonyl ID"
sheet1['C1'] = "Amine ID"
sheet1['D1'] = "Lab Journal Code"
sheet1['E1'] = "Activity Mean"
sheet1['F1'] = "Activity SD"
sheet1['G1'] = "Assay Method"
sheet1['H1'] = "Hit"
sheet1['I1'] = "Combination"

book.save('output/alldata_3SD.xlsx')

file_list = glob("*.xlsx")


def write_data_to_sheet(counter, carbcol, amincol, labcol, actcol, stdcol, assaycol, hitcol, condcol, carbonyl, amine, labcode, activity, std, method, hit, x):
    """
    Writes data to the specified columns in the sheet.

    Args:
        counter (int): The row index where the data should be written.
        carbcol (int): The column index for the carbonyl data.
        amincol (int): The column index for the amine data.
        labcol (int): The column index for the lab code data.
        actcol (int): The column index for the activity data.
        stdcol (int): The column index for the standard data.
        assaycol (int): The column index for the assay method data.
        hitcol (int): The column index for the hit data.
        condcol (int): The column index for the condition data.
        carbonyl (str): The value to be written in the carbonyl column.
        amine (str): The value to be written in the amine column.
        labcode (str): The value to be written in the lab code column.
        activity (str): The value to be written in the activity column.
        std (str): The value to be written in the standard column.
        method (str): The value to be written in the assay method column.
        hit (str): The value to be written in the hit column.
        x (str): The value to be written in the condition column.
    """
    sheet1.cell(row=counter, column=carbcol, value=carbonyl)
    sheet1.cell(row=counter, column=amincol, value=amine)
    sheet1.cell(row=counter, column=labcol, value=labcode)
    sheet1.cell(row=counter, column=actcol, value=activity)
    sheet1.cell(row=counter, column=stdcol, value=std)
    sheet1.cell(row=counter, column=assaycol, value=method)
    sheet1.cell(row=counter, column=hitcol, value=hit)
    sheet1.cell(row=counter, column=condcol, value=x)


# AA="AA" # for 5 SD
AA = "AB"  # uncomment for 3 SD and comment the line above
plate_id = str()


def remodeldata(plate_id, counter):
    """
    Remodels the data from multiple Excel files and writes the transformed data to a new Excel file.

    Args:
        plate_id (str): The ID of the plate to extract data from.
        counter (int): The starting value for the counter.

    Returns:
        int: The updated value of the counter.

    Raises:
        None
    """
    for file in list(file_list):
        data = pd.read_excel(file, sheet_name=None)
        wb = openpyxl.load_workbook(file, data_only=True)
        plate = wb[plate_id]
        enzyme = plate['E3'].value
        labcode = plate['E7'].value
        method = plate['E4'].value
        x = []
        rownum = 11
        actrow = 16
        enzymecol = 1
        carbcol = 2
        amincol = 3
        labcol = 4
        actcol = 5
        stdcol = 6
        assaycol = 7
        hitcol = 8
        condcol = 9
        while rownum < 18:
            sheet1.cell(row=counter, column=enzymecol, value=enzyme)
            rownum = rownum + 1
            B = "B"
            row = str(rownum)
            cell = str(B + row)
            x = plate[cell].value
            carbonyl, amine = x.split("–")
            Y = "Y"
            Z = "Z"
            act = str(actrow)
            cell = str(Y + act)
            cell2 = str(Z + act)
            cell3 = str(AA + act)
            activity = plate[cell].value
            std = plate[cell2].value
            hit = plate[cell3].value
            actrow = actrow + 1
            write_data_to_sheet(counter, carbcol, amincol, labcol, actcol, stdcol, assaycol, hitcol, condcol, carbonyl,
                                amine, labcode, activity, std, method, hit, x)
            counter = counter + 1

        rownum = 11
        actrow = 24
        while rownum < 18:
            sheet1.cell(row=counter, column=enzymecol, value=enzyme)
            rownum = rownum + 1
            E = "E"
            row = str(rownum)
            cell = str(E + row)
            x = plate[cell].value
            carbonyl, amine = x.split("–")
            Y = "Y"
            Z = "Z"
            act = str(actrow)
            cell = str(Y + act)
            cell2 = str(Z + act)
            cell3 = str(AA + act)
            activity = plate[cell].value
            std = plate[cell2].value
            hit = plate[cell3].value
            actrow = actrow + 1
            write_data_to_sheet(counter, carbcol, amincol, labcol, actcol, stdcol, assaycol, hitcol, condcol, carbonyl,
                                amine, labcode, activity, std, method, hit, x)
            counter = counter + 1

        rownum = 11
        actrow = 32
        while rownum < 18:
            sheet1.cell(row=counter, column=enzymecol, value=enzyme)
            rownum = rownum + 1
            H = "H"
            row = str(rownum)
            cell = str(H + row)
            x = plate[cell].value
            carbonyl, amine = x.split("–")
            Y = "Y"
            Z = "Z"
            act = str(actrow)
            cell = str(Y + act)
            cell2 = str(Z + act)
            cell3 = str(AA + act)
            activity = plate[cell].value
            std = plate[cell2].value
            hit = plate[cell3].value
            actrow = actrow + 1
            write_data_to_sheet(counter, carbcol, amincol, labcol, actcol, stdcol, assaycol, hitcol, condcol, carbonyl,
                                amine, labcode, activity, std, method, hit, x)
            counter = counter + 1

        rownum = 11
        actrow = 40
        while rownum < 18:
            sheet1.cell(row=counter, column=enzymecol, value=enzyme)
            rownum = rownum + 1
            K = "K"
            row = str(rownum)
            cell = str(K + row)
            x = plate[cell].value
            carbonyl, amine = x.split("–")
            Y = "Y"
            Z = "Z"
            act = str(actrow)
            cell = str(Y + act)
            cell2 = str(Z + act)
            cell3 = str(AA + act)
            activity = plate[cell].value
            std = plate[cell2].value
            hit = plate[cell3].value
            actrow = actrow + 1
            write_data_to_sheet(counter, carbcol, amincol, labcol, actcol, stdcol, assaycol, hitcol, condcol, carbonyl,
                                amine, labcode, activity, std, method, hit, x)
            counter = counter + 1
            book.save('output/alldata_3SD.xlsx')
    return counter


counter = remodeldata("Plate 1", 2)
remodeldata("Plate 2", counter)

# ----------------------------- isolate hits ----------------------------------------
df = pd.read_excel('output/alldata_3SD.xlsx')

# uncomment to set no-hits to zero instead of removing them entirly
#df['Activity Mean'][pd.isna(df['Hit'])] = 0
#df['Activity SD'][pd.isna(df['Hit'])] = 0

# Filter rows based on 'Hit' column
df = df[df['Hit'].isin(['Hit'])]

# Remove specific combinations from 'Combination' column
df = df[~df['Combination'].isin(['C122–A030', 'C122–A036'])]

# Drop 'Hit' column
df.drop('Hit', axis=1, inplace=True)

# Rename columns
df.columns = ["Enzyme ID", "Carbonyl ID", "Amine ID", "Lab Journal Code", "Activity Mean", "Activity SD",
              "Assay Method", "Combination"]

# Save filtered data to Excel file
df.to_excel('output/matrix3SD-hits-only.xlsx', header=True, index=False)
