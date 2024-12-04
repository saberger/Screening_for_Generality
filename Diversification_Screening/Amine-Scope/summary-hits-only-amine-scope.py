import pandas as pd
from glob import glob
import openpyxl

# Generate environment
workbook = openpyxl.Workbook()
sheet = workbook.active
outname = "output/alldata_amine-scope.xlsx"
workbook.save(outname)

book_ro = openpyxl.load_workbook(outname)
book = openpyxl.Workbook()
book._active_sheet_index = 0
sheet1 = book.active

colx = 1
wb = openpyxl.load_workbook(outname, read_only=True)
sheet = wb.active
counter = 1

sheet1.cell(1, 1, "Enzyme ID")
sheet1.cell(1, 2, "Carbonyl ID")
sheet1.cell(1, 3, "Amine ID")
sheet1.cell(1, 4, "Lab Journal Code")
sheet1.cell(1, 5, "Activity [mU/mgCFE]")
sheet1.cell(1, 6, "Assay Method")
sheet1.cell(1, 7, "Hit")
sheet1.cell(1, 8, "Combination")

book.save(outname)

file_list = glob("*.xlsx")

for file in list(file_list):
    data = pd.read_excel(file, sheet_name=None)
    wb = openpyxl.load_workbook(file, data_only=True)
    plate = wb["Amine Plate"]
    enzyme = plate['E3'].value
    labcode = plate['E7'].value
    method = plate['E4'].value
    x = []
    enzymecol = 0
    carbcol = 1
    amincol = 2
    labcol = 3
    actcol = 4
    assaycol = 5
    hitcol = 6
    condcol = 7
    rownum = 16

    while rownum < 110:
        sheet1.cell(row=counter+1, column=enzymecol+1, value=enzyme)
        row = str(rownum)
        cell = str('O' + row)
        x = plate[cell].value
        carbonyl, amine = x.split("–")
        act = str(rownum)
        cell = str('W' + act)
        cell2 = str('Y' + act)
        activity = plate[cell].value
        hit = plate[cell2].value
        sheet1.cell(row=counter+1, column=carbcol+1, value=carbonyl)
        sheet1.cell(row=counter+1, column=amincol+1, value=amine)
        sheet1.cell(row=counter+1, column=labcol+1, value=labcode)
        sheet1.cell(row=counter+1, column=actcol+1, value=activity)
        sheet1.cell(row=counter+1, column=assaycol+1, value=method)
        sheet1.cell(row=counter+1, column=hitcol+1, value=hit)
        sheet1.cell(row=counter+1, column=condcol+1, value=x)
        counter += 1
        rownum += 1

    book.save(outname)

# Isolate hits
df = pd.read_excel(outname)

combinations_to_drop = ['C105–A029', 'C105–A032', 'C105–A033', 'C139–A029', 'C139–A032', 'C139–A033']
df = df[~df['Combination'].isin(combinations_to_drop)]

#df = df[df['Hit'] == 'Hit']
#df.drop('Hit', axis=1, inplace=True) uncomment to remove non-hits
df['Activity [mU/mgCFE]'][pd.isna(df['Hit'])] = 0


df.to_excel('output/matrix-amine-hits-only.xlsx', header=False, index=False)
df.columns = ["Enzyme ID", "Carbonyl ID", "Amine ID", "Lab Journal Code", "Activity [mU/mgCFE]", "Assay Method", "Combination"]
df.to_excel('output/matrix-amine-hits-only.xlsx', index=False)
