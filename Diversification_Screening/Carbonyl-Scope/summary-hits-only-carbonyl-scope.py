import pandas as pd
from glob import glob
import openpyxl
# --------------------------- combine evaluation sheets ----------------------
workbook = openpyxl.Workbook()
sheet = workbook.active
workbook.save('output/alldata_carbonyl.xlsx')

book_ro = openpyxl.load_workbook('output/alldata_carbonyl.xlsx')
book = openpyxl.Workbook()
book._active_sheet_index = 0
sheet1 = book.active

colx = 1
wb = openpyxl.load_workbook('output/alldata_carbonyl.xlsx', read_only=True)
sheet = wb.active
counter = 1

sheet1.cell(1, 1, "Enzyme ID")
sheet1.cell(1, 2, "Carbonyl ID")
sheet1.cell(1, 3, "Amine ID")
sheet1.cell(1, 4, "Lab Journal Code")
sheet1.cell(1, 5, "Target/Buffer")
sheet1.cell(1, 6, "Target/Buffer")
sheet1.cell(1, 7, "Assay Method")
sheet1.cell(1, 8, "Hit")
sheet1.cell(1, 9, "Combination")

book.save('output/alldata_carbonyl.xlsx')

file_list = glob("*.xlsx")

for file in list(file_list):
    data = pd.read_excel(file, sheet_name=None)
    wb = openpyxl.load_workbook(file, data_only=True)
    plate = wb["Carbonyl Scope"]
    enzyme = plate['B3'].value
    labcode = plate['B7'].value
    method = plate['B4'].value
    x = []
    rownum = 13
    enzymecol = 0
    carbcol = 1
    amincol = 2
    labcol = 3
    actcol = 4
    sn = 5
    assaycol = 6
    hitcol = 7
    condcol = 8

    while rownum < 109:
        sheet1.cell(counter + 1, enzymecol + 1, enzyme)
        rownum += 1
        A = "A"
        row = str(rownum)
        cell = str(A + row)
        x = plate[cell].value
        carbonyl, amine = x.split("–")
        act = str(rownum)
        cell = str("F" + act)
        cell2 = str("L" + act)
        cell3 = str("K" + act)
        activity = plate[cell].value
        hit = plate[cell2].value
        signaltonoise = plate[cell3].value
        sheet1.cell(counter + 1, carbcol + 1, carbonyl)
        sheet1.cell(counter + 1, amincol + 1, amine)
        sheet1.cell(counter + 1, labcol + 1, labcode)
        sheet1.cell(counter + 1, actcol + 1, activity)
        sheet1.cell(counter + 1, sn + 1, signaltonoise)
        sheet1.cell(counter + 1, assaycol + 1, method)
        sheet1.cell(counter + 1, hitcol + 1, hit)
        sheet1.cell(counter + 1, condcol + 1, x)
        counter += 1

    book.save('output/alldata_carbonyl.xlsx')

# ----------------------- isolate hits ------------------------------
df = pd.read_excel('output/alldata_carbonyl.xlsx')
to_drop = ['C078–A002']
df = df[~df['Combination'].isin(to_drop)]
to_drop = ['Hit']
df = df[df['Hit'].isin(to_drop)]
df.drop('Hit', axis=1, inplace=True)
df.columns = ["Enzyme ID", "Carbonyl ID", "Amine ID", "Lab Journal Code",
              "Target/Buffer", "S/N", "Assay Method", "Combination"]
df.to_excel('output/matrix-hits-only.xlsx', index=False)